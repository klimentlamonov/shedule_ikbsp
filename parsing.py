import openpyxl
from os import path


def pars_main(*nmbs):
    """Parsing Excel.
    Function collects data from .xlsx file and return it as a list of dictionary.

    • nmbs - list with numbers of cell where the name of group located.

    • res - list with dictionaries.
            view: [dict, dict, dict, ..., dict]

    • d - dictionary with group schedule data's. Every list there collect four type of information.
        1: Subject name;
        2: Subject type;
        3: Professor name;
        4: Audience;
            view: {'Group'(group name): str,
                    11(monday):{
                        1(first lesson): {
                            111(odd week): list,
                            112(even week): list
                        },
                        2(second lesson): {
                            111(odd week): list,
                            112(even week): list
                        },
                                ...
                        6(sixth lesson): {
                            111(odd week): list,
                            112(even week): list
                        },
                    }
                    12(tuesday):{...}
                    13(wednesday):{...}
                            ...
                    16(saturday):{...}
            }
    """
    res = []
    wb = openpyxl.load_workbook(path.join('.', 'static', 'schedule', '5.xlsx'))
    sheet = wb.active
    for el in nmbs:
        d = {'Group': None,
             11: {
                 1: {111: [],
                     112: []},
                 2: {111: [],
                     112: []},
                 3: {111: [],
                     112: []},
                 4: {111: [],
                     112: []},
                 5: {111: [],
                     112: []},
                 6: {111: [],
                     112: []}},
             12: {
                 1: {111: [],
                     112: []},
                 2: {111: [],
                     112: []},
                 3: {111: [],
                     112: []},
                 4: {111: [],
                     112: []},
                 5: {111: [],
                     112: []},
                 6: {111: [],
                     112: []}},
             13: {
                 1: {111: [],
                     112: []},
                 2: {111: [],
                     112: []},
                 3: {111: [],
                     112: []},
                 4: {111: [],
                     112: []},
                 5: {111: [],
                     112: []},
                 6: {111: [],
                     112: []}},
             14: {
                 1: {111: [],
                     112: []},
                 2: {111: [],
                     112: []},
                 3: {111: [],
                     112: []},
                 4: {111: [],
                     112: []},
                 5: {111: [],
                     112: []},
                 6: {111: [],
                     112: []}},
             15: {
                 1: {111: [],
                     112: []},
                 2: {111: [],
                     112: []},
                 3: {111: [],
                     112: []},
                 4: {111: [],
                     112: []},
                 5: {111: [],
                     112: []},
                 6: {111: [],
                     112: []}},
             16: {
                 1: {111: [],
                     112: []},
                 2: {111: [],
                     112: []},
                 3: {111: [],
                     112: []},
                 4: {111: [],
                     112: []},
                 5: {111: [],
                     112: []},
                 6: {111: [],
                     112: []}},
             }
        count_nmb_weeks = 111
        count_weeks = 11
        count_lessons = 1
        xy = openpyxl.utils.coordinate_to_tuple(el)
        print('Collecting data for ' + sheet[el].value + '... ')
        d['Group'] = sheet[el].value
        for row in sheet[
                   openpyxl.utils.get_column_letter(xy[1]) + str(xy[0] + 2):
                   openpyxl.utils.get_column_letter(xy[1] + 3) + str(xy[0] + 73)]:
            for cell in row:
                d[count_weeks][count_lessons][count_nmb_weeks].append(cell.value)
            if count_nmb_weeks == 111:
                count_nmb_weeks += 1
            else:
                count_nmb_weeks -= 1
                count_lessons += 1
            if count_lessons > 6:
                count_weeks += 1
                count_lessons = 1
        res.append(d)
        print('Done', end='\n\n')
    return res


if __name__ == "__main__":
    print(pars_main('F2', 'K2', 'P2', 'Z2', 'AE2', 'AJ2', 'AT2', 'AY2'))
